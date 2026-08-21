import re
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


EXCEL_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet"
)

_ILLEGAL_EXCEL_CHARACTERS = re.compile(
    r"[\x00-\x08\x0B\x0C\x0E-\x1F]"
)


def excel_safe_text(value: Any) -> str:
    text = _ILLEGAL_EXCEL_CHARACTERS.sub("", str(value or ""))

    # Не позволяем пользовательскому тексту превратиться
    # в формулу после открытия файла в Excel.
    if text.startswith("="):
        return f"'{text}"

    return text


def reviews_export_filename(event: Any) -> str:
    event_code = re.sub(
        r"[^A-Za-z0-9_-]+",
        "_",
        str(event.event_code or event.id),
    ).strip("_")

    return f"reviews_{event_code or event.id}.xlsx"


def build_reviews_workbook(
    event: Any,
    reviews: list[dict[str, Any]],
    status_label: str,
) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отзывы"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A10"

    dark_fill = PatternFill("solid", fgColor="12181A")
    accent_fill = PatternFill("solid", fgColor="E7F4EC")
    accent_font_color = "146B44"
    muted_font_color = "6C7278"
    border_color = "E3E7E5"
    thin_border = Border(
        bottom=Side(style="thin", color=border_color)
    )

    sheet.merge_cells("A1:I1")
    sheet["A1"] = "ОТЗЫВЫ О МЕРОПРИЯТИИ"
    sheet["A1"].fill = dark_fill
    sheet["A1"].font = Font(
        name="Aptos Display",
        size=11,
        bold=True,
        color="FFFFFF",
    )
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells("A2:I2")
    sheet["A2"] = excel_safe_text(event.title)
    sheet["A2"].fill = accent_fill
    sheet["A2"].font = Font(
        name="Aptos Display",
        size=18,
        bold=True,
        color=accent_font_color,
    )
    sheet["A2"].alignment = Alignment(
        vertical="center",
        wrap_text=True,
    )
    sheet.row_dimensions[2].height = 38

    summary_labels = (
        ("A4:B4", "Код мероприятия"),
        ("D4:E4", "Статус"),
        ("G4:H4", "Дата и время"),
        ("A6:B6", "Место"),
        ("D6:E6", "Количество отзывов"),
        ("G6:H6", "Средняя оценка"),
    )

    for cell_range, label in summary_labels:
        sheet.merge_cells(cell_range)
        cell = sheet[cell_range.split(":")[0]]
        cell.value = label
        cell.font = Font(
            name="Aptos",
            size=9,
            bold=True,
            color=muted_font_color,
        )

    for cell_range in (
        "A5:B5",
        "D5:E5",
        "G5:H5",
        "A7:B7",
        "D7:E7",
        "G7:H7",
    ):
        sheet.merge_cells(cell_range)

    sheet["A5"] = excel_safe_text(event.event_code or "Без кода")
    sheet["D5"] = excel_safe_text(status_label)
    sheet["G5"] = datetime.combine(
        event.event_date,
        event.start_time or datetime.min.time(),
    )
    sheet["G5"].number_format = (
        "dd.mm.yyyy hh:mm"
        if event.start_time
        else "dd.mm.yyyy"
    )
    sheet["A7"] = excel_safe_text(event.place or "Не указано")

    first_data_row = 10
    last_data_row = first_data_row + len(reviews) - 1

    sheet["D7"] = (
        f"=COUNTA(A{first_data_row}:A{last_data_row})"
        if reviews
        else "=0"
    )
    sheet["G7"] = (
        f'=IFERROR(AVERAGE(G{first_data_row}:G{last_data_row}),"—")'
        if reviews
        else '="—"'
    )
    sheet["G7"].number_format = "0.0"

    for cell_address in ("A5", "D5", "G5", "A7"):
        sheet[cell_address].font = Font(name="Aptos", size=10)

    for cell_address in ("D7", "G7"):
        sheet[cell_address].font = Font(
            name="Aptos Display",
            size=16,
            bold=True,
            color=accent_font_color,
        )

    headers = [
        "№",
        "KZN ID",
        "Пользователь",
        "Username",
        "ВУЗ",
        "Telegram ID",
        "Оценка",
        "Отзыв",
        "Дата отзыва",
    ]

    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=9, column=column, value=header)
        cell.fill = dark_fill
        cell.font = Font(
            name="Aptos",
            size=10,
            bold=True,
            color="FFFFFF",
        )
        cell.alignment = Alignment(vertical="center")

    sheet.row_dimensions[9].height = 28

    for row_number, review in enumerate(reviews, start=first_data_row):
        values = [
            row_number - first_data_row + 1,
            excel_safe_text(review.get("user_code") or ""),
            excel_safe_text(review.get("full_name") or ""),
            excel_safe_text(
                f"@{review['username']}"
                if review.get("username")
                else ""
            ),
            excel_safe_text(review.get("university") or ""),
            excel_safe_text(review.get("telegram_id") or ""),
            review.get("rating"),
            excel_safe_text(
                review.get("text") or "Без комментария"
            ),
            review.get("created_at"),
        ]

        for column, value in enumerate(values, start=1):
            cell = sheet.cell(
                row=row_number,
                column=column,
                value=value,
            )
            cell.font = Font(name="Aptos", size=10)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(column in {3, 5, 8}),
            )
            cell.border = thin_border

        sheet.cell(row=row_number, column=7).number_format = "0"

        created_at = sheet.cell(row=row_number, column=9)
        if isinstance(created_at.value, datetime):
            created_at.number_format = "dd.mm.yyyy hh:mm"

        sheet.row_dimensions[row_number].height = 42

    if reviews:
        table = Table(
            displayName=f"EventReviews{event.id}",
            ref=f"A9:I{last_data_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

        sheet.conditional_formatting.add(
            f"G{first_data_row}:G{last_data_row}",
            ColorScaleRule(
                start_type="num",
                start_value=1,
                start_color="F8696B",
                mid_type="num",
                mid_value=3,
                mid_color="FFEB84",
                end_type="num",
                end_value=5,
                end_color="63BE7B",
            ),
        )
    else:
        sheet.merge_cells("A10:I11")
        sheet["A10"] = "Отзывов на мероприятие пока нет."
        sheet["A10"].font = Font(
            name="Aptos",
            size=11,
            italic=True,
            color=muted_font_color,
        )
        sheet["A10"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    column_widths = {
        "A": 7,
        "B": 16,
        "C": 29,
        "D": 20,
        "E": 24,
        "F": 17,
        "G": 11,
        "H": 58,
        "I": 20,
    }

    for column_letter, width in column_widths.items():
        sheet.column_dimensions[column_letter].width = width

    sheet.print_title_rows = "1:9"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_orientation = "landscape"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
