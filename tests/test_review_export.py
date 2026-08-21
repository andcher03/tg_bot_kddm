from datetime import date, datetime, time
from io import BytesIO
from types import SimpleNamespace
from zipfile import ZipFile

from openpyxl import load_workbook

from web_admin.review_export import (
    build_reviews_workbook,
    excel_safe_text,
    reviews_export_filename,
)


def make_event():
    return SimpleNamespace(
        id=4,
        event_code="EVENT-000004",
        title="Тестовое мероприятие",
        event_date=date(2026, 8, 23),
        start_time=time(18, 0),
        place="Казань",
    )


def test_build_reviews_workbook():
    event = make_event()
    reviews = [
        {
            "user_code": "KZN-000007",
            "full_name": "Иван Иванов",
            "username": "ivanov",
            "university": "КФУ",
            "telegram_id": 123456789,
            "rating": 5,
            "text": "Отличное мероприятие",
            "created_at": datetime(2026, 8, 24, 12, 30),
        },
        {
            "user_code": "KZN-000008",
            "full_name": "Тест формулы",
            "username": None,
            "university": None,
            "telegram_id": 987654321,
            "rating": 3,
            "text": "=2+2",
            "created_at": datetime(2026, 8, 24, 13, 0),
        },
    ]

    output = build_reviews_workbook(
        event,
        reviews,
        "Актуально",
    )
    workbook = load_workbook(output, data_only=False)
    sheet = workbook["Отзывы"]

    assert sheet["A2"].value == event.title
    assert sheet["D7"].value == "=COUNTA(A10:A11)"
    assert sheet["G7"].value == '=IFERROR(AVERAGE(G10:G11),"—")'
    assert sheet["G10"].value == 5
    assert sheet["H11"].value == "'=2+2"
    assert sheet["I10"].value == reviews[0]["created_at"]
    assert sheet["I10"].number_format == "dd.mm.yyyy hh:mm"
    assert sheet.freeze_panes == "A10"
    assert len(sheet.tables) == 1

    with ZipFile(BytesIO(output.getvalue())) as archive:
        sheet_xml = archive.read(
            "xl/worksheets/sheet1.xml"
        ).decode("utf-8")
        table_xml = archive.read(
            "xl/tables/table1.xml"
        ).decode("utf-8")
        workbook_xml = archive.read(
            "xl/workbook.xml"
        ).decode("utf-8")

    # Фильтр должен принадлежать только Excel-таблице.
    # Дублирование фильтра на уровне листа повреждает книгу в Excel.
    assert "<autoFilter" not in sheet_xml
    assert table_xml.count("<autoFilter") == 1
    assert "_xlnm._FilterDatabase" not in workbook_xml


def test_export_helpers():
    assert excel_safe_text("=HYPERLINK(\"https://example.com\")").startswith("'")
    assert reviews_export_filename(make_event()) == (
        "reviews_EVENT-000004.xlsx"
    )


def test_empty_reviews_workbook_has_clear_empty_state():
    output = build_reviews_workbook(
        make_event(),
        [],
        "Актуально",
    )
    workbook = load_workbook(output, data_only=False)
    sheet = workbook["Отзывы"]

    assert sheet["D7"].value == "=0"
    assert sheet["G7"].value == '="—"'
    assert sheet["A10"].value == "Отзывов на мероприятие пока нет."
    assert len(sheet.tables) == 0
